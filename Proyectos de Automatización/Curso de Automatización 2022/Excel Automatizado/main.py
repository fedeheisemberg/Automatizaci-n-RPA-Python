import pandas as pd
from openpyxl import load_workbook
import openpyxl
from openpyxl.chart import BarChart, Reference
import string
from openpyxl.styles import Font

def automatizar_excel(nombre_archivo):

    """Input sales_mes.xlsx/ Output repor_mes.xlsx"""
    archivo_excel=pd.read_excel('supermarket_sales.xlsx')


    tabla_pivote=pd.pivot_table(data=archivo_excel,index='Gender',columns='Product line',values='Total',aggfunc='sum')

    mes_extension=nombre_archivo.split('_')[1]
    tabla_pivote.to_excel(f'sales_{mes_extension}.xlsx',startrow=4,sheet_name='Report')

    wb = load_workbook('sales_2021.xlsx')
    pestaña = wb['Report']

    min_columna=wb.active.min_column
    max_columna=wb.active.max_column
    min_fila=wb.active.min_row
    max_fila=wb.active.max_row


    #Gráfico

    barchart=BarChart()

    data=Reference(pestaña,min_col=min_columna+1,min_row=min_fila+1,max_row=max_fila,max_col=max_columna)
    categorias=Reference(pestaña,min_col=min_columna,min_row=min_fila+1,max_row=max_fila)

    barchart.add_data(data,titles_from_data=True)
    barchart.set_categories(categorias)

    pestaña.add_chart(barchart,'B12')
    barchart.title='Ventas'
    barchart.style=2

    abecedario=list(string.ascii_uppercase)
    abecedario_excel=abecedario[0:max_columna]

    for i in abecedario_excel:
        if i!='A':
            pestaña[f'{i}{max_fila+1}']=f'=SUM({i}{min_fila+1}:{i}{min_fila+1})'
            pestaña[f'{i}{max_fila+1}'].style='Currency'

    pestaña[f'{abecedario_excel[0]}{max_fila+1}']='Total'
    pestaña['A1']='Reporte'
    mes=mes_extension.split('.')[0]
    pestaña['A2']=mes
    pestaña['A1'].font=Font('Arial',bold=True,size=20)
    pestaña['A2'].font=Font('Arial',bold=True,size=12)

    wb.save(f'sales_{mes_extension}.xlsx')
    return

#automatizar reporte 2021
automatizar_excel('sales_2021.xlsx')

#automatizar reporte 2021